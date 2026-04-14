import io
import re
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

COLUNAS = [
    "TipoDocumento","Numero","Cliente","Empresa","CNPJ",
    "Data","Valor","ChaveAcesso","Banco","Pagador",
    "Recebedor","Autenticacao","Observacoes"
]

TIPOS = {
    "nota_fiscal":"Nota Fiscal","boleto":"Boleto",
    "comprovante_pix":"PIX","comprovante_ted":"TED",
    "comprovante":"Comprovante","cheque":"Cheque",
    "recibo":"Recibo","desconhecido":"Desconhecido"
}

def _limpar_cnpj(valor):
    if not valor: return ""
    n = re.sub(r"\D", "", str(valor))
    if len(n) == 14: return n
    m = re.search(r"\d{14}", n)
    return m.group(0) if m else ""

def _limpar_cliente(valor):
    if not valor: return ""
    valor = re.sub(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}", "", valor)
    valor = re.sub(r"\d{3}\.\d{3}\.\d{3}-\d{2}", "", valor)
    valor = re.sub(r"\d{2}/\d{2}/\d{2,4}", "", valor)
    valor = re.sub(r"\s\d{6,}\s?", " ", valor)
    return re.sub(r"\s+", " ", valor).strip()

def _limpar_empresa(valor):
    if not valor: return ""
    v_upper = valor.upper()
    if any(x in v_upper for x in ["ORTHOMAX","COLCHOES","DAVI"]):
        return "ORTHOMAX INDUSTRIA E COMERCIO DE COLCHOES"
    valor = re.sub(r"^[^A-Za-z]+", "", valor)
    valor = re.sub(r"Cond Pag.*$", "", valor, flags=re.IGNORECASE)
    valor = re.sub(r"\b(DANFE|DOCUMENTO|AUXILIAR|DA NOTA|FISCAL)\b", "", valor, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", valor).strip()[:60]

def _limpar_valor(valor):
    if not valor: return ""
    s = str(valor).replace(",",".")
    partes = s.split(".")
    if len(partes) > 2:
        s = "".join(partes[:-1]) + "." + partes[-1]
    try: return str(round(float(s), 2))
    except: return ""

def _limpar_chave(valor):
    if not valor: return ""
    n = re.sub(r"\D", "", str(valor))
    if len(n) == 44: return n
    m = re.search(r"\d{44}", n)
    return m.group(0) if m else ""

def _mapear(doc_type, fields):
    row = {}
    row["TipoDocumento"] = TIPOS.get(doc_type, doc_type)
    numero_raw = (fields.get("numero_nota") or fields.get("numero_recibo") or
                  fields.get("numero_cheque") or fields.get("numero_pedido") or "")
    n = re.sub(r"\D", "", str(numero_raw))
    row["Numero"] = n if n and len(n) <= 10 else ""
    cliente_raw = (fields.get("cliente") or fields.get("destinatario") or
                   fields.get("pagador") or fields.get("beneficiario") or "")
    row["Cliente"] = _limpar_cliente(cliente_raw)[:60]
    empresa_raw = (fields.get("empresa") or fields.get("nome_empresa") or
                   fields.get("empresa_emissora") or "")
    row["Empresa"] = _limpar_empresa(empresa_raw)
    cnpj_raw = (fields.get("cnpj_emissor") or fields.get("cnpj") or
                fields.get("cnpj_cliente") or fields.get("cnpj_beneficiario") or
                fields.get("cnpj_cpf") or "")
    row["CNPJ"] = _limpar_cnpj(cnpj_raw)
    row["Data"] = (fields.get("data_emissao") or fields.get("data") or
                   fields.get("vencimento") or "")
    row["Valor"] = _limpar_valor(fields.get("valor_total") or fields.get("valor") or "")
    row["ChaveAcesso"] = _limpar_chave(fields.get("chave_acesso") or fields.get("linha_digitavel") or "")
    row["Banco"] = fields.get("banco") or fields.get("banco_emissor") or ""
    row["Pagador"] = _limpar_cliente(fields.get("pagador") or "")[:50]
    row["Recebedor"] = _limpar_cliente(fields.get("recebedor") or fields.get("beneficiario") or "")[:50]
    row["Autenticacao"] = fields.get("autenticacao") or fields.get("protocolo") or ""
    obs = []
    if fields.get("serie"): obs.append("Serie: "+str(fields["serie"]))
    if fields.get("natureza_operacao"): obs.append(str(fields["natureza_operacao"])[:40])
    if fields.get("chave_pix"): obs.append("PIX: "+str(fields["chave_pix"]))
    row["Observacoes"] = " | ".join(obs)
    return row

def _validar(row):
    if not row.get("TipoDocumento"): return False, "sem TipoDocumento"
    if not row.get("Numero"): return False, "sem Numero"
    uteis = sum([bool(row.get("Cliente")), bool(row.get("Empresa")),
                 bool(row.get("CNPJ")), bool(row.get("Data")), bool(row.get("Valor"))])
    if uteis < 2: return False, f"poucos campos ({uteis})"
    return True, "ok"

def _deduplicar(rows):
    vistos = {}
    resultado = []
    for row in rows:
        chave = (row["TipoDocumento"], row["Numero"])
        if chave in vistos and row["Numero"]:
            idx = vistos[chave]
            for col in COLUNAS:
                if not resultado[idx].get(col) and row.get(col):
                    resultado[idx][col] = row[col]
        else:
            vistos[chave] = len(resultado)
            resultado.append(row)
    return resultado

def _preparar_linhas(data, doc_type):
    documentos = data.get("documentos", [])
    if not documentos:
        fields = data.get("fields", {})
        if fields:
            documentos = [{"doc_type": doc_type, "fields": fields}]
    rows_mapeadas = []
    for doc in documentos:
        tipo = doc.get("doc_type", doc_type)
        fields = doc.get("fields", {})
        if not fields: continue
        row = _mapear(tipo, fields)
        rows_mapeadas.append(row)
    rows_dedup = _deduplicar(rows_mapeadas)
    rows_validas = []
    for row in rows_dedup:
        ok, motivo = _validar(row)
        if ok:
            rows_validas.append(row)
            print("OK:", row.get("TipoDocumento"), row.get("Numero"))
        else:
            print("IGNORADO:", motivo, row.get("TipoDocumento"), row.get("Numero"))
    return rows_validas, documentos

def write_to_excel(file_bytes, data, doc_type):
    print("=== write_to_excel ===")

    # Se usuario enviou planilha, usar o formato dela
    colunas_usar = COLUNAS
    if file_bytes and len(file_bytes) > 100:
        try:
            from openpyxl import load_workbook as _lw
            wb_user = _lw(filename=io.BytesIO(file_bytes))
            ws_user = wb_user.active
            cab_user = [c.value for c in ws_user[1] if c.value]
            if cab_user and len(cab_user) >= 3:
                colunas_usar = cab_user
                print("Usando formato da planilha do usuario:", colunas_usar)
                wb = wb_user
                ws = ws_user
                # Encontrar proxima linha vazia
                prox_linha = ws_user.max_row + 1
                for r in range(2, ws_user.max_row + 1):
                    if all(ws_user.cell(r, c).value is None for c in range(1, len(cab_user)+1)):
                        prox_linha = r
                        break
            else:
                raise ValueError("Cabecalho invalido")
        except Exception as e:
            print("Erro ao ler planilha do usuario, criando nova:", e)
            wb = Workbook()
            ws = wb.active
            ws.title = "DocFinance"
            prox_linha = 2
            for col_num, col_name in enumerate(COLUNAS, 1):
                cell = ws.cell(row=1, column=col_num, value=col_name)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="4F46E5")
                cell.alignment = Alignment(horizontal="center")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "DocFinance"
        prox_linha = 2
        for col_num, col_name in enumerate(COLUNAS, 1):
            cell = ws.cell(row=1, column=col_num, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F46E5")
            cell.alignment = Alignment(horizontal="center")

    rows_validas, documentos = _preparar_linhas(data, doc_type)
    print(f"Documentos brutos: {len(documentos)}")
    print(f"Linhas validas: {len(rows_validas)}")
    print(f"Inserindo a partir da linha: {prox_linha}")

    for i, row_data in enumerate(rows_validas):
        linha_num = prox_linha + i
        cor = "F0F0FF" if linha_num % 2 == 0 else "FFFFFF"
        for col_num, col_name in enumerate(colunas_usar, 1):
            val = row_data.get(col_name) or None
            cell = ws.cell(row=linha_num, column=col_num, value=val)
            try:
                cell.fill = PatternFill("solid", fgColor=cor)
                cell.alignment = Alignment(vertical="center")
            except: pass

    larguras = {"TipoDocumento":15,"Numero":12,"Cliente":30,"Empresa":35,
                "CNPJ":18,"Data":12,"Valor":14,"ChaveAcesso":46,"Banco":15,
                "Pagador":25,"Recebedor":25,"Autenticacao":20,"Observacoes":40}
    for i, col in enumerate(colunas_usar, 1):
        try:
            ws.column_dimensions[ws.cell(1,i).column_letter].width = larguras.get(col, 20)
        except: pass

    # ABA ITENS
    ws_itens = wb.create_sheet("Itens")
    cols_itens = ["NumeroNota","Codigo","Descricao","Quantidade","ValorUnitario","ValorTotalItem"]
    for col_num, col_name in enumerate(cols_itens, 1):
        cell = ws_itens.cell(row=1, column=col_num, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="10B981")
        cell.alignment = Alignment(horizontal="center")
    linha_item = 2
    for doc in documentos:
        tipo = doc.get("doc_type", doc_type)
        fields = doc.get("fields", {})
        if not fields: continue
        row_principal = _mapear(tipo, fields)
        num_nota = row_principal.get("Numero", "")
        for item in fields.get("itens", []):
            ws_itens.cell(row=linha_item, column=1, value=num_nota)
            ws_itens.cell(row=linha_item, column=2, value=item.get("codigo",""))
            ws_itens.cell(row=linha_item, column=3, value=item.get("descricao",""))
            ws_itens.cell(row=linha_item, column=4, value=item.get("quantidade",""))
            ws_itens.cell(row=linha_item, column=5, value=item.get("valor_unitario",""))
            ws_itens.cell(row=linha_item, column=6, value=item.get("valor_total_item",""))
            linha_item += 1
    ws_itens.column_dimensions["A"].width = 12
    ws_itens.column_dimensions["B"].width = 10
    ws_itens.column_dimensions["C"].width = 50
    ws_itens.column_dimensions["D"].width = 12
    ws_itens.column_dimensions["E"].width = 15
    ws_itens.column_dimensions["F"].width = 15
    print(f"Planilha salva com {len(rows_validas)} linhas + {linha_item-2} itens")
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def write_to_google_sheets(spreadsheet_id, data, doc_type):
    import gspread
    from google.oauth2.service_account import Credentials
    import os

    print("=== write_to_google_sheets ===")
    print("Spreadsheet ID:", spreadsheet_id)

    # Autenticar
    creds_path = os.path.join(os.path.dirname(__file__), "..", "credentials.json")
    if not os.path.exists(creds_path):
        raise FileNotFoundError("credentials.json nao encontrado em: " + creds_path)

    scopes = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive"
    ]
    creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
    gc = gspread.authorize(creds)
    print("Autenticado como service account")

    # Abrir planilha pelo ID
    try:
        sh = gc.open_by_key(spreadsheet_id)
        print("Planilha aberta:", sh.title)
        print("Abas disponiveis:", [ws.title for ws in sh.worksheets()])
    except Exception as e:
        raise Exception("Erro ao abrir planilha: " + str(e))

    # Preparar linhas
    rows_validas, documentos = _preparar_linhas(data, doc_type)
    print("Total de registros para enviar:", len(rows_validas))

    # === ABA RESUMO ===
    cab_resumo = ["TipoDocumento","Numero","Cliente","Empresa","CNPJ","Data","Valor","Observacoes"]
    try:
        ws_resumo = sh.worksheet("Resumo")
        print("Aba selecionada: Resumo")
    except gspread.WorksheetNotFound:
        ws_resumo = sh.add_worksheet("Resumo", rows=1000, cols=len(cab_resumo))
        print("Aba Resumo criada")

    # UPSERT robusto por chave TipoDocumento+Numero
    existing = ws_resumo.get_all_values()
    print("Registros recebidos para processar:", len(rows_validas))
    print("Linhas existentes na aba Resumo (incluindo cabecalho):", len(existing))

    # Descobrir posicao das colunas no cabecalho REAL da planilha
    cabecalho_real = existing[0] if existing else cab_resumo
    try:
        idx_tipo = cabecalho_real.index("TipoDocumento")
        idx_num = cabecalho_real.index("Numero")
    except ValueError:
        # Se cabecalho nao tem essas colunas, usar posicoes padrao
        idx_tipo = 0
        idx_num = 1

    # Montar indice por chave
    chaves_existentes = {}
    for i, row in enumerate(existing[1:], start=2):
        if len(row) > max(idx_tipo, idx_num):
            tipo_val = row[idx_tipo].strip()
            num_val = row[idx_num].strip()
            if tipo_val or num_val:
                chave = (tipo_val, num_val)
                chaves_existentes[chave] = i

    print("Chaves ja existentes na planilha:", list(chaves_existentes.keys()))

    linhas_novas = []
    atualizadas = 0

    for row_data in rows_validas:
        linha = [str(row_data.get(col, "") or "") for col in cab_resumo]
        chave = (linha[idx_tipo].strip(), linha[idx_num].strip())

        if chave[1] and chave in chaves_existentes:
            # Atualizar linha existente
            linha_num = chaves_existentes[chave]
            col_final = chr(64 + len(cab_resumo))
            range_str = f"A{linha_num}:{col_final}{linha_num}"
            ws_resumo.update(range_str, [linha], value_input_option="USER_ENTERED")
            print(f"Atualizado: {chave} na linha {linha_num}")
            atualizadas += 1
        else:
            linhas_novas.append(linha)
            print(f"Novo registro: {chave}")

    if linhas_novas:
        ws_resumo.append_rows(linhas_novas, value_input_option="USER_ENTERED")

    print(f"Resumo final: {atualizadas} atualizados, {len(linhas_novas)} inseridos")

    # === ABA ITENS ===
    cab_itens = ["NumeroNota","Codigo","Descricao","Quantidade","ValorUnitario","ValorTotalItem"]
    try:
        ws_itens = sh.worksheet("Itens")
        print("Aba selecionada: Itens")
    except gspread.WorksheetNotFound:
        ws_itens = sh.add_worksheet("Itens", rows=5000, cols=len(cab_itens))
        print("Aba Itens criada")

    # Limpar dados antigos mantendo cabecalho
    existing_itens = ws_itens.get_all_values()
    print("Linhas existentes na aba Itens antes de limpar:", len(existing_itens))
    ws_itens.clear()
    ws_itens.append_row(cab_itens)
    print("Aba Itens limpa - cabecalho mantido na linha 1")

    linhas_itens = []
    for doc in documentos:
        tipo = doc.get("doc_type", doc_type)
        fields = doc.get("fields", {})
        if not fields:
            continue
        row_principal = _mapear(tipo, fields)
        num_nota = row_principal.get("Numero", "")
        for item in fields.get("itens", []):
            linhas_itens.append([
                str(num_nota),
                str(item.get("codigo", "")),
                str(item.get("descricao", "")),
                str(item.get("quantidade", "")),
                str(item.get("valor_unitario", "")),
                str(item.get("valor_total_item", ""))
            ])

    print("Linhas a inserir na aba Itens:", len(linhas_itens))
    if linhas_itens:
        ws_itens.append_rows(linhas_itens, value_input_option="USER_ENTERED")
        print("Linhas inseridas com sucesso na aba Itens!")

    return True


def list_spreadsheets():
    import gspread
    from google.oauth2.service_account import Credentials
    import os
    creds_path = os.path.join(os.path.dirname(__file__), "..", "credentials.json")
    if not os.path.exists(creds_path):
        return []
    try:
        scopes = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
        gc = gspread.authorize(creds)
        return [{"id": s.id, "name": s.title} for s in gc.openall()]
    except Exception as e:
        print("Erro list_spreadsheets:", e)
        return []

def create_spreadsheet(title="DocFinance"):
    import gspread
    from google.oauth2.service_account import Credentials
    import os
    creds_path = os.path.join(os.path.dirname(__file__), "..", "credentials.json")
    scopes = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
    gc = gspread.authorize(creds)
    sh = gc.create(title)
    sh.share(None, perm_type="anyone", role="writer")
    return sh.id
