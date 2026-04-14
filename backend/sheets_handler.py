"""
Integração com Google Sheets e Excel para DocFinance.
"""

import os
import logging
from pathlib import Path
from typing import List, Dict, Optional, Any
from datetime import datetime

logger = logging.getLogger("docfinance.sheets")

OUTPUTS_DIR = Path("outputs")
OUTPUTS_DIR.mkdir(exist_ok=True)

# ── Excel ────────────────────────────────────────────────────────────────────

COLUMN_HEADERS = {
    "boleto": [
        "Arquivo", "Tipo Documento", "Linha Digitável", "Beneficiário",
        "Pagador", "CNPJ/CPF", "Valor (R$)", "Data Vencimento",
        "Banco", "Nosso Número", "Confiança OCR", "Data Processamento"
    ],
    "nota_fiscal": [
        "Arquivo", "Tipo Documento", "Número Nota", "CNPJ Emissor",
        "Nome Empresa", "Data Emissão", "Valor Total (R$)", "ICMS (R$)",
        "ISS (R$)", "Confiança OCR", "Data Processamento"
    ],
    "comprovante": [
        "Arquivo", "Tipo Documento", "Tipo Transação", "Valor (R$)",
        "Data", "Hora", "Pagador", "Recebedor", "Banco",
        "Código Autenticação", "Confiança OCR", "Data Processamento"
    ],
    "cheque": [
        "Arquivo", "Tipo Documento", "Número Cheque", "Beneficiário",
        "Valor (R$)", "Valor Por Extenso", "Data", "Banco",
        "Cidade", "Confiança OCR", "Data Processamento"
    ],
    "desconhecido": [
        "Arquivo", "Tipo Documento", "Texto Extraído", "Confiança OCR", "Data Processamento"
    ]
}

FIELD_MAP = {
    "boleto": {
        "arquivo": "Arquivo", "tipo_documento": "Tipo Documento",
        "linha_digitavel": "Linha Digitável", "beneficiario": "Beneficiário",
        "pagador": "Pagador", "cnpj_cpf": "CNPJ/CPF", "valor": "Valor (R$)",
        "data_vencimento": "Data Vencimento", "banco": "Banco",
        "nosso_numero": "Nosso Número", "confianca_ocr": "Confiança OCR",
        "data_processamento": "Data Processamento"
    },
    "nota_fiscal": {
        "arquivo": "Arquivo", "tipo_documento": "Tipo Documento",
        "numero_nota": "Número Nota", "cnpj_emissor": "CNPJ Emissor",
        "nome_empresa": "Nome Empresa", "data_emissao": "Data Emissão",
        "valor_total": "Valor Total (R$)", "icms": "ICMS (R$)",
        "iss": "ISS (R$)", "confianca_ocr": "Confiança OCR",
        "data_processamento": "Data Processamento"
    },
    "comprovante": {
        "arquivo": "Arquivo", "tipo_documento": "Tipo Documento",
        "tipo": "Tipo Transação", "valor": "Valor (R$)",
        "data": "Data", "hora": "Hora", "pagador": "Pagador",
        "recebedor": "Recebedor", "banco": "Banco",
        "codigo_autenticacao": "Código Autenticação", "confianca_ocr": "Confiança OCR",
        "data_processamento": "Data Processamento"
    },
    "cheque": {
        "arquivo": "Arquivo", "tipo_documento": "Tipo Documento",
        "numero_cheque": "Número Cheque", "beneficiario": "Beneficiário",
        "valor": "Valor (R$)", "valor_extenso": "Valor Por Extenso",
        "data": "Data", "banco": "Banco", "cidade": "Cidade",
        "confianca_ocr": "Confiança OCR", "data_processamento": "Data Processamento"
    }
}


async def write_to_excel(job_id: str, results: List[Dict], sheet_name: str = "Extrações") -> str:
    """
    Escreve resultados em arquivo Excel organizado por tipo de documento.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove aba padrão
        
        # Agrupar por tipo de documento
        by_type: Dict[str, List] = {}
        for result in results:
            if result.get("status") == "erro":
                by_type.setdefault("erros", []).append(result)
                continue
            doc_type = result.get("data", {}).get("tipo_documento", "desconhecido")
            by_type.setdefault(doc_type, []).append(result)
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1a3a5c", end_color="1a3a5c", fill_type="solid")
        revisar_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        na_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC")
        )
        
        type_names = {
            "boleto": "Boletos",
            "nota_fiscal": "Notas Fiscais",
            "comprovante": "Comprovantes",
            "cheque": "Cheques",
            "desconhecido": "Não Identificado",
            "erros": "Erros"
        }
        
        colors = {
            "boleto": "2563EB",
            "nota_fiscal": "16A34A",
            "comprovante": "7C3AED",
            "cheque": "D97706",
            "desconhecido": "6B7280",
            "erros": "DC2626"
        }
        
        for doc_type, type_results in by_type.items():
            ws = wb.create_sheet(title=type_names.get(doc_type, doc_type)[:31])
            
            headers = COLUMN_HEADERS.get(doc_type, ["Arquivo", "Informação"])
            field_map = FIELD_MAP.get(doc_type, {})
            
            # Header row
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.fill = PatternFill(
                    start_color=colors.get(doc_type, "1a3a5c"),
                    end_color=colors.get(doc_type, "1a3a5c"),
                    fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            
            ws.row_dimensions[1].height = 30
            
            # Dados
            for row_idx, result in enumerate(type_results, 2):
                if doc_type == "erros":
                    ws.cell(row=row_idx, column=1, value=result.get("file", ""))
                    ws.cell(row=row_idx, column=2, value=result.get("error", "Erro desconhecido"))
                    continue
                
                campos = result.get("data", {}).get("campos", {})
                campos_revisao = result.get("data", {}).get("campos_revisao", [])
                
                for col_idx, header in enumerate(headers, 1):
                    # Encontrar campo correspondente
                    field_key = next((k for k, v in field_map.items() if v == header), None)
                    value = campos.get(field_key, "N/A") if field_key else "N/A"
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=value if value is not None else "N/A")
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")
                    
                    # Destacar campos que precisam revisão
                    if field_key and field_key in campos_revisao:
                        cell.fill = revisar_fill
                        cell.font = Font(bold=True, color="856404")
                    elif str(value) == "N/A":
                        cell.fill = na_fill
                        cell.font = Font(color="6B7280")
                
                ws.row_dimensions[row_idx].height = 22
            
            # Auto-ajustar colunas
            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)
            
            # Congelar linha do cabeçalho
            ws.freeze_panes = "A2"
        
        # Aba de resumo
        if len(wb.sheetnames) > 0:
            ws_summary = wb.create_sheet(title="Resumo", index=0)
            ws_summary["A1"] = "DocFinance - Resumo do Processamento"
            ws_summary["A1"].font = Font(bold=True, size=14, color="1a3a5c")
            ws_summary["A3"] = "Tipo de Documento"
            ws_summary["B3"] = "Quantidade"
            ws_summary["C3"] = "Status"
            
            for cell in ws_summary[3]:
                if cell.value:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1a3a5c", end_color="1a3a5c", fill_type="solid")
            
            row = 4
            total = 0
            for doc_type, type_results in by_type.items():
                ws_summary.cell(row=row, column=1, value=type_names.get(doc_type, doc_type))
                ws_summary.cell(row=row, column=2, value=len(type_results))
                ws_summary.cell(row=row, column=3, value="✓ Processado" if doc_type != "erros" else "✗ Erro")
                total += len(type_results)
                row += 1
            
            ws_summary.cell(row=row+1, column=1, value="TOTAL").font = Font(bold=True)
            ws_summary.cell(row=row+1, column=2, value=total).font = Font(bold=True)
            ws_summary[f"A{row+3}"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws_summary[f"A{row+3}"].font = Font(italic=True, color="6B7280")
            
            ws_summary.column_dimensions["A"].width = 25
            ws_summary.column_dimensions["B"].width = 15
            ws_summary.column_dimensions["C"].width = 20
        
        # Salvar
        output_path = OUTPUTS_DIR / f"docfinance_{job_id[:8]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(str(output_path))
        logger.info(f"Excel salvo: {output_path}")
        return str(output_path)
    
    except ImportError:
        logger.error("openpyxl não instalado!")
        raise RuntimeError("openpyxl é necessário para gerar Excel")
    except Exception as e:
        logger.error(f"Erro ao gerar Excel: {e}")
        raise


# ── Google Sheets ────────────────────────────────────────────────────────────

GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID", "")
GOOGLE_CLIENT_SECRET = os.getenv("GOOGLE_CLIENT_SECRET", "")
GOOGLE_REDIRECT_URI = os.getenv("GOOGLE_REDIRECT_URI", "http://localhost:8000/google/callback")
GOOGLE_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.readonly"
]


def get_google_auth_url(user_id: str) -> str:
    """Gera URL de autenticação OAuth Google."""
    if not GOOGLE_CLIENT_ID:
        return "GOOGLE_CLIENT_ID não configurado. Configure no .env"
    
    try:
        from google_auth_oauthlib.flow import Flow
        flow = Flow.from_client_config(
            {
                "web": {
                    "client_id": GOOGLE_CLIENT_ID,
                    "client_secret": GOOGLE_CLIENT_SECRET,
                    "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                    "token_uri": "https://oauth2.googleapis.com/token",
                    "redirect_uris": [GOOGLE_REDIRECT_URI]
                }
            },
            scopes=GOOGLE_SCOPES
        )
        flow.redirect_uri = GOOGLE_REDIRECT_URI
        auth_url, _ = flow.authorization_url(
            access_type="offline",
            include_granted_scopes="true",
            state=user_id
        )
        return auth_url
    except ImportError:
        return "google-auth-oauthlib não instalado"


async def exchange_google_code(code: str, state: str) -> Dict:
    """Troca código OAuth por tokens."""
    from storage import google_tokens
    try:
        from google_auth_oauthlib.flow import Flow
        flow = Flow.from_client_config(
            {
                "web": {
                    "client_id": GOOGLE_CLIENT_ID,
                    "client_secret": GOOGLE_CLIENT_SECRET,
                    "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                    "token_uri": "https://oauth2.googleapis.com/token",
                    "redirect_uris": [GOOGLE_REDIRECT_URI]
                }
            },
            scopes=GOOGLE_SCOPES,
            state=state
        )
        flow.redirect_uri = GOOGLE_REDIRECT_URI
        flow.fetch_token(code=code)
        
        google_tokens[state] = {
            "token": flow.credentials.token,
            "refresh_token": flow.credentials.refresh_token,
            "token_uri": flow.credentials.token_uri,
            "client_id": flow.credentials.client_id,
            "client_secret": flow.credentials.client_secret,
            "scopes": list(flow.credentials.scopes or [])
        }
        return {"message": "Google Sheets conectado com sucesso!"}
    except Exception as e:
        logger.error(f"Erro OAuth Google: {e}")
        return {"error": str(e)}


async def list_user_sheets(user_id: str) -> List[Dict]:
    """Lista planilhas Google do usuário."""
    from storage import google_tokens
    creds_data = google_tokens.get(user_id)
    if not creds_data:
        return []
    
    try:
        from google.oauth2.credentials import Credentials
        from googleapiclient.discovery import build
        
        creds = Credentials(**creds_data)
        service = build("drive", "v3", credentials=creds)
        
        results = service.files().list(
            q="mimeType='application/vnd.google-apps.spreadsheet'",
            fields="files(id, name, modifiedTime)",
            pageSize=20
        ).execute()
        
        return results.get("files", [])
    except Exception as e:
        logger.error(f"Erro ao listar planilhas: {e}")
        return []


async def write_to_google_sheets(sheet_id: str, results: List[Dict], sheet_name: str = "Extrações") -> str:
    """Escreve dados no Google Sheets."""
    from storage import google_tokens
    
    try:
        from google.oauth2.credentials import Credentials
        from googleapiclient.discovery import build
        
        # Encontrar credenciais (simplificado - em produção: buscar por sheet_id)
        creds_data = next(iter(google_tokens.values()), None)
        if not creds_data:
            raise RuntimeError("Usuário não conectou Google Sheets")
        
        creds = Credentials(**creds_data)
        service = build("sheets", "v4", credentials=creds)
        
        # Preparar dados
        all_rows = [["Arquivo", "Tipo", "Campo", "Valor", "Status", "Data Processamento"]]
        
        for result in results:
            if result.get("status") == "erro":
                all_rows.append([result.get("file"), "ERRO", "-", result.get("error"), "Erro", datetime.now().isoformat()])
                continue
            
            campos = result.get("data", {}).get("campos", {})
            campos_revisao = result.get("data", {}).get("campos_revisao", [])
            
            for campo, valor in campos.items():
                if campo in ("arquivo", "tipo_documento", "data_processamento", "confianca_ocr"):
                    continue
                status = "REVISAR" if campo in campos_revisao else ("N/A" if valor == "N/A" else "OK")
                all_rows.append([
                    campos.get("arquivo", ""),
                    campos.get("tipo_documento", ""),
                    campo,
                    str(valor),
                    status,
                    campos.get("data_processamento", "")
                ])
        
        # Escrever em lote
        body = {"values": all_rows}
        service.spreadsheets().values().append(
            spreadsheetId=sheet_id,
            range=f"{sheet_name}!A1",
            valueInputOption="USER_ENTERED",
            insertDataOption="INSERT_ROWS",
            body=body
        ).execute()
        
        logger.info(f"Dados escritos no Google Sheets: {sheet_id}")
        return f"https://docs.google.com/spreadsheets/d/{sheet_id}"
    
    except Exception as e:
        logger.error(f"Erro ao escrever no Google Sheets: {e}")
        raise
